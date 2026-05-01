"""
Full test suite for app.py logic functions.

Run with coverage:
    pytest --cov=app --cov-report=term-missing --cov-report=html
"""
import json
import pandas as pd
import pytest
import streamlit as st

import app
from app import (
    _add_to_group,
    _delete_group,
    _find_group,
    _merge_groups,
    _sync_field,
    apply_cross_batch_merge,
    build_cross_batch_prompt,
    build_prompt,
    ensure_group_ids,
    load_history,
    load_results_record,
    merge_all_batches,
    normalise_quotes,
    parse_response,
    preprocess,
    save_monthly_data,
)


# ── Shared factory ────────────────────────────────────────────────────────────

def make_group(
    gid="g1",
    app_name="TestApp",
    issue="Test issue",
    ids=None,
    impact="Some impact",
    action="Some action",
):
    ids = ids or ["INC001"]
    return {
        "_id": gid,
        "application": app_name,
        "issue": issue,
        "incident_numbers": list(ids),
        "count": len(ids),
        "business_impact": impact,
        "recommended_action": action,
    }


# ══════════════════════════════════════════════════════════════════════════════
# preprocess
# ══════════════════════════════════════════════════════════════════════════════

class TestPreprocess:
    def test_removes_ticket_id_pattern(self):
        assert "INC12345" not in preprocess("Fix INC12345 urgently")

    def test_removes_chg_ticket(self):
        assert "CHG99999" not in preprocess("Related to CHG99999")

    def test_removes_email(self):
        assert "user@example.com" not in preprocess("Contact user@example.com")

    def test_removes_tel_prefix(self):
        assert "01234567890" not in preprocess("Call Tel: 01234567890")

    def test_removes_international_phone(self):
        result = preprocess("Call +44 7700 900 123")
        assert "+44" not in result

    def test_does_not_remove_short_numbers(self):
        # short numbers like port 8080 or error code 404 should survive
        result = preprocess("Error 404 on port 8080")
        assert "404" in result
        assert "8080" in result

    def test_removes_cid_disclaimer(self):
        assert "Client Identifying Data" not in preprocess(
            "Client Identifying Data not allowed here."
        )

    def test_removes_no_cid_flag(self):
        result = preprocess("No CID Disclaimer accepted: TRUE")
        assert "CID" not in result.upper() or "DISCLAIMER" not in result.upper()

    def test_removes_read_disclaimer_line(self):
        assert "disclaimer" not in preprocess(
            "I have read and understood the disclaimer and agree."
        ).lower()

    def test_removes_hi_greeting(self):
        result = preprocess("Hi, I cannot log in to Bloomberg")
        assert not result.lower().startswith("hi")

    def test_removes_hello_greeting(self):
        result = preprocess("Hello, please help with access")
        assert not result.lower().startswith("hello")

    def test_removes_dear_greeting(self):
        result = preprocess("Dear support team, we have an issue")
        assert not result.lower().startswith("dear")

    def test_removes_thanks_signoff(self):
        result = preprocess("Please fix the login. Thanks")
        assert "thanks" not in result.lower()

    def test_removes_regards_signoff(self):
        result = preprocess("Need access restored. Best regards")
        assert "regards" not in result.lower()

    def test_removes_cheers_signoff(self):
        result = preprocess("Sorting this would help. Cheers")
        assert "cheers" not in result.lower()

    def test_removes_user_details_block(self):
        result = preprocess("User details: John Smith Topic: Access issue")
        assert "User details" not in result

    def test_preserves_topic_after_user_details(self):
        result = preprocess("User details: John Smith Topic: Access issue")
        assert "Access issue" in result

    def test_removes_t_ref_token(self):
        assert "(T12345)" not in preprocess("Ticket (T12345) was raised")

    def test_collapses_double_dash_separator(self):
        assert "---" not in preprocess("Subject --- body")

    def test_collapses_underscore_separator(self):
        assert "___" not in preprocess("Section _____ end")

    def test_collapses_extra_whitespace(self):
        result = preprocess("too   many    spaces")
        assert "  " not in result

    def test_truncates_to_custom_maxlen(self):
        assert len(preprocess("a" * 1000, maxlen=100)) == 100

    def test_default_maxlen_is_500(self):
        assert len(preprocess("x" * 600)) == 500

    def test_coerces_non_string_to_string(self):
        assert isinstance(preprocess(42), str)

    def test_empty_string_returns_empty(self):
        assert preprocess("") == ""

    def test_preserves_meaningful_content(self):
        text = "Bloomberg terminal crashes on startup when loading the portfolio view"
        result = preprocess(text)
        assert "Bloomberg" in result
        assert "crash" in result


# ══════════════════════════════════════════════════════════════════════════════
# normalise_quotes
# ══════════════════════════════════════════════════════════════════════════════

class TestNormaliseQuotes:
    def test_left_double_quote(self):
        assert normalise_quotes("“word”") == '"word"'

    def test_right_double_quote(self):
        assert normalise_quotes("say ”hello") == 'say "hello'

    def test_left_single_quote(self):
        assert normalise_quotes("it’s broken") == "it's broken"

    def test_right_single_quote(self):
        assert normalise_quotes("don‘t") == "don't"

    def test_prime_single(self):
        assert normalise_quotes("5′") == "5'"

    def test_prime_double(self):
        assert normalise_quotes('12″') == '12"'

    def test_strips_bom(self):
        assert normalise_quotes("﻿hello") == "hello"

    def test_plain_ascii_unchanged(self):
        assert normalise_quotes("plain text 123") == "plain text 123"

    def test_multiple_replacements_in_one_string(self):
        result = normalise_quotes("“hello” it’s ‘fine’")
        assert result == '"hello" it\'s \'fine\''


# ══════════════════════════════════════════════════════════════════════════════
# ensure_group_ids
# ══════════════════════════════════════════════════════════════════════════════

class TestEnsureGroupIds:
    def test_adds_id_to_group_without_one(self):
        groups = [{"issue": "test"}]
        ensure_group_ids(groups)
        assert "_id" in groups[0]

    def test_id_is_nonempty_string(self):
        groups = [{"issue": "test"}]
        ensure_group_ids(groups)
        assert isinstance(groups[0]["_id"], str)
        assert len(groups[0]["_id"]) > 0

    def test_preserves_existing_id(self):
        groups = [{"_id": "keep-me", "issue": "test"}]
        ensure_group_ids(groups)
        assert groups[0]["_id"] == "keep-me"

    def test_empty_list_returns_empty(self):
        assert ensure_group_ids([]) == []

    def test_each_group_gets_unique_id(self):
        groups = [{"issue": "a"}, {"issue": "b"}, {"issue": "c"}]
        ensure_group_ids(groups)
        ids = [g["_id"] for g in groups]
        assert len(set(ids)) == 3

    def test_mixed_with_and_without_ids(self):
        groups = [{"_id": "existing"}, {"issue": "no id yet"}]
        ensure_group_ids(groups)
        assert groups[0]["_id"] == "existing"
        assert "_id" in groups[1]
        assert groups[1]["_id"] != "existing"

    def test_returns_same_list_object(self):
        groups = [{"issue": "x"}]
        result = ensure_group_ids(groups)
        assert result is groups


# ══════════════════════════════════════════════════════════════════════════════
# parse_response
# ══════════════════════════════════════════════════════════════════════════════

class TestParseResponse:

    def _wrap(self, groups_list):
        return json.dumps({"groups": groups_list})

    def _group(self, app="App", issue="Issue", ids=None, impact="", action=""):
        return {
            "application": app,
            "issue": issue,
            "incident_numbers": ids or ["INC001"],
            "business_impact": impact,
            "recommended_action": action,
        }

    def test_parses_valid_json(self):
        result = parse_response(self._wrap([self._group()]))
        assert result is not None
        assert len(result) == 1
        assert result[0]["application"] == "App"

    def test_strips_markdown_json_fence(self):
        raw = "```json\n" + self._wrap([self._group()]) + "\n```"
        assert parse_response(raw) is not None

    def test_strips_plain_fence(self):
        raw = "```\n" + self._wrap([self._group()]) + "\n```"
        assert parse_response(raw) is not None

    def test_returns_none_on_no_json(self):
        assert parse_response("no braces here at all") is None

    def test_returns_none_on_invalid_json(self):
        assert parse_response("{ this is not : valid json }") is None

    def test_deduplicates_ids_within_group(self):
        g = self._group(ids=["INC001", "INC001", "INC002"])
        result = parse_response(self._wrap([g]))
        assert result[0]["incident_numbers"] == ["INC001", "INC002"]

    def test_deduplicates_ids_across_groups(self):
        g1 = self._group(issue="A", ids=["INC001"])
        g2 = self._group(issue="B", ids=["INC001", "INC002"])
        result = parse_response(self._wrap([g1, g2]))
        all_ids = [n for grp in result for n in grp["incident_numbers"]]
        assert all_ids.count("INC001") == 1

    def test_skips_group_with_no_unique_ids(self):
        g1 = self._group(issue="A", ids=["INC001"])
        g2 = self._group(issue="B", ids=["INC001"])  # INC001 already seen
        result = parse_response(self._wrap([g1, g2]))
        assert len(result) == 1

    def test_strips_whitespace_from_ids(self):
        g = self._group(ids=[" INC001 ", "INC002"])
        result = parse_response(self._wrap([g]))
        assert "INC001" in result[0]["incident_numbers"]
        assert " INC001 " not in result[0]["incident_numbers"]

    def test_count_reflects_unique_ids(self):
        g = self._group(ids=["INC001", "INC002"])
        result = parse_response(self._wrap([g]))
        assert result[0]["count"] == 2

    def test_assigns_uuid_to_each_group(self):
        g1 = self._group(issue="A", ids=["INC001"])
        g2 = self._group(issue="B", ids=["INC002"])
        result = parse_response(self._wrap([g1, g2]))
        assert "_id" in result[0]
        assert "_id" in result[1]
        assert result[0]["_id"] != result[1]["_id"]

    def test_falls_back_to_unknown_system(self):
        g = {"issue": "Test", "incident_numbers": ["INC001"],
             "business_impact": "", "recommended_action": ""}
        result = parse_response(self._wrap([g]))
        assert result[0]["application"] == "Unknown System"

    def test_accepts_alternative_key_g(self):
        raw = json.dumps({"g": [self._group()]})
        result = parse_response(raw)
        assert result is not None

    def test_empty_groups_list(self):
        result = parse_response(self._wrap([]))
        assert result == []

    def test_normalises_curly_quotes_in_json(self):
        # Wrapping with curly double-quotes that normalise_quotes converts
        raw = '“groups”: [{"application": "App", "issue": "x", "incident_numbers": ["INC001"], "business_impact": "", "recommended_action": ""}]}'
        raw = "{" + raw
        result = parse_response(raw)
        assert result is not None

    def test_scratchpad_text_before_json_is_ignored(self):
        preamble = "Here are my thoughts on the groupings...\n\n"
        raw = preamble + self._wrap([self._group()])
        result = parse_response(raw)
        assert result is not None
        assert len(result) == 1


# ══════════════════════════════════════════════════════════════════════════════
# apply_cross_batch_merge
# ══════════════════════════════════════════════════════════════════════════════

class TestApplyCrossBatchMerge:

    def _make_original(self, n):
        return [
            {
                "_id": f"id{i}",
                "application": f"App{i}",
                "issue": f"Issue {i}",
                "incident_numbers": [f"INC{i:03d}"],
                "count": 1,
                "business_impact": "",
                "recommended_action": "",
            }
            for i in range(1, n + 1)
        ]

    def _spec(self, app, issue, indices):
        return {
            "application": app,
            "issue": issue,
            "business_impact": "",
            "recommended_action": "",
            "source_indices": indices,
        }

    def test_merges_two_groups(self):
        original = self._make_original(2)
        result = apply_cross_batch_merge(original, [self._spec("App", "Merged", [1, 2])])
        assert result is not None
        assert len(result) == 1
        assert set(result[0]["incident_numbers"]) == {"INC001", "INC002"}
        assert result[0]["count"] == 2

    def test_keeps_single_group_separate(self):
        original = self._make_original(1)
        result = apply_cross_batch_merge(original, [self._spec("App1", "A", [1])])
        assert result is not None
        assert len(result) == 1
        assert result[0]["incident_numbers"] == ["INC001"]

    def test_mix_of_merged_and_separate(self):
        original = self._make_original(3)
        spec = [
            self._spec("Merged", "AB", [1, 2]),
            self._spec("App3",   "C",  [3]),
        ]
        result = apply_cross_batch_merge(original, spec)
        assert result is not None
        assert len(result) == 2
        merged = next(r for r in result if len(r["incident_numbers"]) > 1)
        assert set(merged["incident_numbers"]) == {"INC001", "INC002"}

    def test_returns_none_when_not_all_assigned(self):
        original = self._make_original(3)
        # group 3 not mentioned
        result = apply_cross_batch_merge(original, [self._spec("App", "AB", [1, 2])])
        assert result is None

    def test_returns_none_on_out_of_range_index(self):
        original = self._make_original(2)
        # index 99 is out of range → zero_based becomes empty → returns None
        result = apply_cross_batch_merge(original, [self._spec("App", "X", [99, 1])])
        # index 1 valid but 99 filtered; effectively only [1] assigned, group 2 unassigned
        assert result is None

    def test_returns_none_on_empty_source_indices(self):
        original = self._make_original(1)
        result = apply_cross_batch_merge(original, [self._spec("App", "X", [])])
        assert result is None

    def test_empty_originals_and_spec(self):
        assert apply_cross_batch_merge([], []) == []

    def test_assigns_new_uuid(self):
        original = self._make_original(2)
        result = apply_cross_batch_merge(original, [self._spec("App", "M", [1, 2])])
        assert "_id" in result[0]

    def test_uses_spec_application_name(self):
        original = self._make_original(2)
        result = apply_cross_batch_merge(original, [self._spec("NewApp", "M", [1, 2])])
        assert result[0]["application"] == "NewApp"


# ══════════════════════════════════════════════════════════════════════════════
# merge_all_batches
# ══════════════════════════════════════════════════════════════════════════════

class TestMergeAllBatches:

    def test_single_batch_single_group(self):
        result = merge_all_batches([{"groups": [make_group(ids=["INC001"])]}])
        assert len(result) == 1

    def test_deduplicates_across_batches(self):
        b1 = {"groups": [make_group(ids=["INC001"])]}
        b2 = {"groups": [make_group(ids=["INC001", "INC002"])]}
        result = merge_all_batches([b1, b2])
        all_ids = [n for g in result for n in g["incident_numbers"]]
        assert all_ids.count("INC001") == 1

    def test_merges_same_app_and_issue(self):
        g1 = make_group(gid="a", app_name="AD", issue="Login fail", ids=["INC001"])
        g2 = make_group(gid="b", app_name="AD", issue="Login fail", ids=["INC002"])
        result = merge_all_batches([{"groups": [g1]}, {"groups": [g2]}])
        assert len(result) == 1
        assert set(result[0]["incident_numbers"]) == {"INC001", "INC002"}

    def test_keeps_distinct_groups_separate(self):
        g1 = make_group(gid="a", app_name="App1", issue="A", ids=["INC001"])
        g2 = make_group(gid="b", app_name="App2", issue="B", ids=["INC002"])
        result = merge_all_batches([{"groups": [g1, g2]}])
        assert len(result) == 2

    def test_case_insensitive_dedup_on_app_and_issue(self):
        g1 = make_group(gid="a", app_name="bloomberg", issue="crash", ids=["INC001"])
        g2 = make_group(gid="b", app_name="Bloomberg", issue="Crash", ids=["INC002"])
        result = merge_all_batches([{"groups": [g1]}, {"groups": [g2]}])
        assert len(result) == 1

    def test_empty_batch_list(self):
        assert merge_all_batches([]) == []

    def test_batch_missing_groups_key(self):
        result = merge_all_batches([{}])
        assert result == []

    def test_assigns_uuid_to_new_groups(self):
        g = make_group(gid="a", ids=["INC001"])
        del g["_id"]  # simulate group without _id
        result = merge_all_batches([{"groups": [g]}])
        assert "_id" in result[0]

    def test_count_updated_after_merge(self):
        g1 = make_group(gid="a", app_name="AD", issue="X", ids=["INC001"])
        g2 = make_group(gid="b", app_name="AD", issue="X", ids=["INC002", "INC003"])
        result = merge_all_batches([{"groups": [g1]}, {"groups": [g2]}])
        assert result[0]["count"] == 3


# ══════════════════════════════════════════════════════════════════════════════
# _find_group
# ══════════════════════════════════════════════════════════════════════════════

class TestFindGroup:

    def test_finds_group_by_id(self):
        g = make_group(gid="abc")
        st.session_state.all_groups = [g]
        assert _find_group("abc") is g

    def test_returns_none_for_unknown_id(self):
        st.session_state.all_groups = [make_group(gid="abc")]
        assert _find_group("xyz") is None

    def test_returns_none_when_all_groups_is_none(self):
        st.session_state.all_groups = None
        assert _find_group("abc") is None

    def test_returns_none_on_empty_list(self):
        st.session_state.all_groups = []
        assert _find_group("abc") is None

    def test_returns_correct_group_among_many(self):
        groups = [make_group(gid=f"g{i}", issue=f"Issue {i}") for i in range(5)]
        st.session_state.all_groups = groups
        result = _find_group("g3")
        assert result["issue"] == "Issue 3"


# ══════════════════════════════════════════════════════════════════════════════
# _sync_field
# ══════════════════════════════════════════════════════════════════════════════

class TestSyncField:

    def test_updates_issue_field(self):
        g = make_group(gid="abc", issue="old issue")
        st.session_state.all_groups = [g]
        st.session_state["issue_abc"] = "new issue"
        _sync_field("abc", "issue", "issue_abc")
        assert g["issue"] == "new issue"

    def test_updates_business_impact(self):
        g = make_group(gid="abc", impact="old")
        st.session_state.all_groups = [g]
        st.session_state["imp_abc"] = "updated impact"
        _sync_field("abc", "business_impact", "imp_abc")
        assert g["business_impact"] == "updated impact"

    def test_updates_recommended_action(self):
        g = make_group(gid="abc", action="old action")
        st.session_state.all_groups = [g]
        st.session_state["act_abc"] = "new action"
        _sync_field("abc", "recommended_action", "act_abc")
        assert g["recommended_action"] == "new action"

    def test_no_crash_when_group_not_found(self):
        st.session_state.all_groups = []
        st.session_state["issue_xyz"] = "value"
        _sync_field("xyz", "issue", "issue_xyz")  # must not raise

    def test_empty_string_value_is_accepted(self):
        g = make_group(gid="abc", issue="had text")
        st.session_state.all_groups = [g]
        st.session_state["issue_abc"] = ""
        _sync_field("abc", "issue", "issue_abc")
        assert g["issue"] == ""


# ══════════════════════════════════════════════════════════════════════════════
# _delete_group
# ══════════════════════════════════════════════════════════════════════════════

class TestDeleteGroup:

    def test_removes_group_from_all_groups(self):
        g = make_group(gid="abc")
        st.session_state.all_groups = [g]
        _delete_group("abc")
        assert len(st.session_state.all_groups) == 0

    def test_moves_incident_numbers_to_missing_ids(self):
        g = make_group(gid="abc", ids=["INC001", "INC002"])
        st.session_state.all_groups = [g]
        st.session_state.missing_ids = []
        _delete_group("abc")
        assert "INC001" in st.session_state.missing_ids
        assert "INC002" in st.session_state.missing_ids

    def test_merges_freed_ids_with_existing_missing(self):
        g = make_group(gid="abc", ids=["INC003"])
        st.session_state.all_groups = [g]
        st.session_state.missing_ids = ["INC004"]
        _delete_group("abc")
        assert "INC003" in st.session_state.missing_ids
        assert "INC004" in st.session_state.missing_ids

    def test_removes_gid_from_merge_selected(self):
        g = make_group(gid="abc")
        st.session_state.all_groups = [g]
        st.session_state.merge_selected = {"abc", "other"}
        _delete_group("abc")
        assert "abc" not in st.session_state.merge_selected
        assert "other" in st.session_state.merge_selected

    def test_keeps_other_groups_intact(self):
        g1 = make_group(gid="abc", ids=["INC001"])
        g2 = make_group(gid="def", ids=["INC002"])
        st.session_state.all_groups = [g1, g2]
        _delete_group("abc")
        assert len(st.session_state.all_groups) == 1
        assert st.session_state.all_groups[0]["_id"] == "def"

    def test_no_crash_when_group_not_found(self):
        st.session_state.all_groups = []
        st.session_state.missing_ids = []
        _delete_group("nonexistent")  # must not raise

    def test_missing_ids_sorted_after_delete(self):
        g = make_group(gid="abc", ids=["INC003", "INC001", "INC002"])
        st.session_state.all_groups = [g]
        st.session_state.missing_ids = []
        _delete_group("abc")
        assert st.session_state.missing_ids == sorted(st.session_state.missing_ids)


# ══════════════════════════════════════════════════════════════════════════════
# _add_to_group
# ══════════════════════════════════════════════════════════════════════════════

class TestAddToGroup:

    def test_adds_new_ids_to_group(self):
        g = make_group(gid="abc", ids=["INC001"])
        st.session_state.all_groups = [g]
        st.session_state.missing_ids = ["INC002"]
        _add_to_group("abc", ["INC002"])
        assert "INC002" in g["incident_numbers"]

    def test_removes_added_ids_from_missing(self):
        g = make_group(gid="abc", ids=["INC001"])
        st.session_state.all_groups = [g]
        st.session_state.missing_ids = ["INC002", "INC003"]
        _add_to_group("abc", ["INC002"])
        assert "INC002" not in st.session_state.missing_ids
        assert "INC003" in st.session_state.missing_ids

    def test_updates_count(self):
        g = make_group(gid="abc", ids=["INC001"])
        st.session_state.all_groups = [g]
        st.session_state.missing_ids = ["INC002", "INC003"]
        _add_to_group("abc", ["INC002", "INC003"])
        assert g["count"] == 3

    def test_deduplicates_if_id_already_in_group(self):
        g = make_group(gid="abc", ids=["INC001"])
        st.session_state.all_groups = [g]
        st.session_state.missing_ids = ["INC001"]
        _add_to_group("abc", ["INC001"])
        assert g["incident_numbers"].count("INC001") == 1

    def test_empty_ids_list_is_noop(self):
        g = make_group(gid="abc", ids=["INC001"])
        st.session_state.all_groups = [g]
        original_count = g["count"]
        _add_to_group("abc", [])
        assert g["count"] == original_count

    def test_no_crash_when_group_not_found(self):
        st.session_state.all_groups = []
        st.session_state.missing_ids = ["INC001"]
        _add_to_group("nope", ["INC001"])  # must not raise

    def test_multiple_ids_added_at_once(self):
        g = make_group(gid="abc", ids=[])
        g["count"] = 0
        st.session_state.all_groups = [g]
        st.session_state.missing_ids = ["INC001", "INC002", "INC003"]
        _add_to_group("abc", ["INC001", "INC002", "INC003"])
        assert g["count"] == 3
        assert st.session_state.missing_ids == []


# ══════════════════════════════════════════════════════════════════════════════
# _merge_groups
# ══════════════════════════════════════════════════════════════════════════════

class TestMergeGroups:

    def test_merges_two_groups_into_one(self):
        g1 = make_group(gid="a", ids=["INC001"])
        g2 = make_group(gid="b", ids=["INC002"])
        st.session_state.all_groups = [g1, g2]
        st.session_state.merge_selected = {"a", "b"}
        _merge_groups(["a", "b"], "New issue")
        assert len(st.session_state.all_groups) == 1

    def test_merged_group_contains_all_incident_numbers(self):
        g1 = make_group(gid="a", ids=["INC001"])
        g2 = make_group(gid="b", ids=["INC002"])
        st.session_state.all_groups = [g1, g2]
        st.session_state.merge_selected = {"a", "b"}
        _merge_groups(["a", "b"], "New issue")
        assert set(st.session_state.all_groups[0]["incident_numbers"]) == {"INC001", "INC002"}

    def test_merged_group_uses_provided_issue(self):
        g1 = make_group(gid="a", ids=["INC001"])
        g2 = make_group(gid="b", ids=["INC002"])
        st.session_state.all_groups = [g1, g2]
        st.session_state.merge_selected = {"a", "b"}
        _merge_groups(["a", "b"], "Custom merged issue")
        assert st.session_state.all_groups[0]["issue"] == "Custom merged issue"

    def test_empty_new_issue_falls_back_to_first_group_issue(self):
        g1 = make_group(gid="a", issue="Original", ids=["INC001"])
        g2 = make_group(gid="b", ids=["INC002"])
        st.session_state.all_groups = [g1, g2]
        st.session_state.merge_selected = {"a", "b"}
        _merge_groups(["a", "b"], "")
        assert st.session_state.all_groups[0]["issue"] == "Original"

    def test_merges_three_groups(self):
        groups = [make_group(gid=f"g{i}", ids=[f"INC00{i}"]) for i in range(1, 4)]
        st.session_state.all_groups = groups
        st.session_state.merge_selected = {"g1", "g2", "g3"}
        _merge_groups(["g1", "g2", "g3"], "Triple merge")
        assert len(st.session_state.all_groups) == 1
        assert st.session_state.all_groups[0]["count"] == 3

    def test_clears_merge_selected_after_merge(self):
        g1 = make_group(gid="a", ids=["INC001"])
        g2 = make_group(gid="b", ids=["INC002"])
        st.session_state.all_groups = [g1, g2]
        st.session_state.merge_selected = {"a", "b"}
        _merge_groups(["a", "b"], "M")
        assert st.session_state.merge_selected == set()

    def test_preserves_unselected_groups(self):
        g1 = make_group(gid="a", ids=["INC001"])
        g2 = make_group(gid="b", ids=["INC002"])
        g3 = make_group(gid="c", issue="Keep me", ids=["INC003"])
        st.session_state.all_groups = [g1, g2, g3]
        st.session_state.merge_selected = {"a", "b"}
        _merge_groups(["a", "b"], "M")
        issues = [g["issue"] for g in st.session_state.all_groups]
        assert "Keep me" in issues

    def test_deduplicates_incident_numbers_across_merged_groups(self):
        g1 = make_group(gid="a", ids=["INC001", "INC002"])
        g2 = make_group(gid="b", ids=["INC002", "INC003"])
        st.session_state.all_groups = [g1, g2]
        st.session_state.merge_selected = {"a", "b"}
        _merge_groups(["a", "b"], "M")
        ids = st.session_state.all_groups[0]["incident_numbers"]
        assert ids.count("INC002") == 1

    def test_count_updated_after_merge(self):
        g1 = make_group(gid="a", ids=["INC001", "INC002"])
        g2 = make_group(gid="b", ids=["INC003"])
        st.session_state.all_groups = [g1, g2]
        st.session_state.merge_selected = {"a", "b"}
        _merge_groups(["a", "b"], "M")
        assert st.session_state.all_groups[0]["count"] == 3

    def test_noop_with_only_one_matching_group(self):
        g = make_group(gid="a", ids=["INC001"])
        st.session_state.all_groups = [g]
        st.session_state.merge_selected = {"a"}
        _merge_groups(["a"], "M")
        assert len(st.session_state.all_groups) == 1

    def test_noop_with_no_matching_gids(self):
        g = make_group(gid="a", ids=["INC001"])
        st.session_state.all_groups = [g]
        st.session_state.merge_selected = set()
        _merge_groups(["x", "y"], "M")
        assert len(st.session_state.all_groups) == 1

    def test_merged_group_gets_new_uuid(self):
        g1 = make_group(gid="a", ids=["INC001"])
        g2 = make_group(gid="b", ids=["INC002"])
        st.session_state.all_groups = [g1, g2]
        st.session_state.merge_selected = {"a", "b"}
        _merge_groups(["a", "b"], "M")
        new_id = st.session_state.all_groups[0]["_id"]
        assert new_id not in ("a", "b")

    def test_inherits_application_from_first_source(self):
        g1 = make_group(gid="a", app_name="AppA", ids=["INC001"])
        g2 = make_group(gid="b", app_name="AppB", ids=["INC002"])
        st.session_state.all_groups = [g1, g2]
        st.session_state.merge_selected = {"a", "b"}
        _merge_groups(["a", "b"], "M")
        assert st.session_state.all_groups[0]["application"] == "AppA"


# ══════════════════════════════════════════════════════════════════════════════
# build_prompt
# ══════════════════════════════════════════════════════════════════════════════

class TestBuildPrompt:

    def test_contains_incident_number(self):
        rows = [{"number": "INC001", "description": "Bloomberg crash"}]
        assert "INC001" in build_prompt(rows)

    def test_contains_description(self):
        rows = [{"number": "INC001", "description": "Bloomberg crash on login"}]
        assert "Bloomberg crash on login" in build_prompt(rows)

    def test_incident_count_in_header(self):
        rows = [{"number": f"INC{i:03d}", "description": "x"} for i in range(7)]
        assert "7" in build_prompt(rows)

    def test_all_incidents_numbered(self):
        rows = [{"number": "INC001", "description": "a"},
                {"number": "INC002", "description": "b"}]
        prompt = build_prompt(rows)
        assert "1." in prompt
        assert "2." in prompt

    def test_contains_json_output_format(self):
        rows = [{"number": "INC001", "description": "x"}]
        assert "groups" in build_prompt(rows)


# ══════════════════════════════════════════════════════════════════════════════
# build_cross_batch_prompt
# ══════════════════════════════════════════════════════════════════════════════

class TestBuildCrossBatchPrompt:

    def test_lists_group_numbers(self):
        groups = [make_group(issue="Login fail")]
        assert "1." in build_cross_batch_prompt(groups)

    def test_includes_application_name(self):
        groups = [make_group(app_name="Bloomberg", issue="Terminal crash")]
        assert "Bloomberg" in build_cross_batch_prompt(groups)

    def test_includes_issue_text(self):
        groups = [make_group(issue="Unique issue description XYZ")]
        assert "Unique issue description XYZ" in build_cross_batch_prompt(groups)

    def test_includes_group_count_in_header(self):
        groups = [make_group(gid=f"g{i}") for i in range(4)]
        assert "4" in build_cross_batch_prompt(groups)

    def test_mentions_source_indices_in_format(self):
        groups = [make_group()]
        assert "source_indices" in build_cross_batch_prompt(groups)


# ══════════════════════════════════════════════════════════════════════════════
# save_monthly_data & load_history
# ══════════════════════════════════════════════════════════════════════════════

class TestSaveMonthlyData:

    def test_creates_csv_file(self, tmp_path, monkeypatch):
        monkeypatch.setattr(app, "DATA_DIR", tmp_path)
        df = pd.DataFrame({"number": ["INC001"], "description_raw": ["desc"]})
        save_monthly_data(df, [], "2025", "January")
        assert (tmp_path / "raw_data_2025_January.csv").exists()

    def test_creates_json_file(self, tmp_path, monkeypatch):
        monkeypatch.setattr(app, "DATA_DIR", tmp_path)
        df = pd.DataFrame({"number": ["INC001"], "description_raw": ["desc"]})
        save_monthly_data(df, [], "2025", "January")
        assert (tmp_path / "processed_data_2025_January.json").exists()

    def test_json_contains_year_month_groups(self, tmp_path, monkeypatch):
        monkeypatch.setattr(app, "DATA_DIR", tmp_path)
        df = pd.DataFrame({"number": ["INC001"], "description_raw": ["desc"]})
        groups = [make_group(ids=["INC001"])]
        save_monthly_data(df, groups, "2025", "March")
        data = json.loads((tmp_path / "processed_data_2025_March.json").read_text())
        assert data["year"] == "2025"
        assert data["month"] == "March"
        assert len(data["groups"]) == 1

    def test_csv_contains_number_and_description_raw(self, tmp_path, monkeypatch):
        monkeypatch.setattr(app, "DATA_DIR", tmp_path)
        df = pd.DataFrame({"number": ["INC001", "INC002"],
                           "description_raw": ["desc1", "desc2"]})
        save_monthly_data(df, [], "2025", "April")
        loaded = pd.read_csv(tmp_path / "raw_data_2025_April.csv")
        assert list(loaded.columns) == ["number", "description_raw"]
        assert len(loaded) == 2

    def test_creates_data_dir_if_missing(self, tmp_path, monkeypatch):
        target = tmp_path / "newdir"
        monkeypatch.setattr(app, "DATA_DIR", target)
        df = pd.DataFrame({"number": ["INC001"], "description_raw": ["d"]})
        save_monthly_data(df, [], "2025", "May")
        assert target.exists()


class TestLoadHistory:

    def test_returns_empty_when_dir_missing(self, tmp_path, monkeypatch):
        monkeypatch.setattr(app, "DATA_DIR", tmp_path / "nonexistent")
        assert load_history() == []

    def test_returns_empty_when_no_json_files(self, tmp_path, monkeypatch):
        monkeypatch.setattr(app, "DATA_DIR", tmp_path)
        assert load_history() == []

    def test_loads_valid_json_file(self, tmp_path, monkeypatch):
        monkeypatch.setattr(app, "DATA_DIR", tmp_path)
        data = {"year": "2025", "month": "January", "groups": []}
        (tmp_path / "processed_data_2025_January.json").write_text(json.dumps(data))
        result = load_history()
        assert len(result) == 1
        assert result[0]["month"] == "January"

    def test_skips_invalid_json(self, tmp_path, monkeypatch):
        monkeypatch.setattr(app, "DATA_DIR", tmp_path)
        (tmp_path / "processed_data_2025_January.json").write_text("not valid json")
        assert load_history() == []

    def test_loads_multiple_files(self, tmp_path, monkeypatch):
        monkeypatch.setattr(app, "DATA_DIR", tmp_path)
        for month in ["January", "February", "March"]:
            data = {"year": "2025", "month": month, "groups": []}
            (tmp_path / f"processed_data_2025_{month}.json").write_text(json.dumps(data))
        assert len(load_history()) == 3

    def test_ignores_non_matching_filenames(self, tmp_path, monkeypatch):
        monkeypatch.setattr(app, "DATA_DIR", tmp_path)
        (tmp_path / "some_other_file.json").write_text(json.dumps({"x": 1}))
        assert load_history() == []

    def test_mixed_valid_and_invalid(self, tmp_path, monkeypatch):
        monkeypatch.setattr(app, "DATA_DIR", tmp_path)
        (tmp_path / "processed_data_2025_January.json").write_text(
            json.dumps({"year": "2025", "month": "January", "groups": []})
        )
        (tmp_path / "processed_data_2025_February.json").write_text("bad json")
        result = load_history()
        assert len(result) == 1


# ══════════════════════════════════════════════════════════════════════════════
# load_results_record
# ══════════════════════════════════════════════════════════════════════════════

class TestLoadResultsRecord:

    def _record(self, ids=None, year="2025", month="January"):
        ids = ids or ["INC001", "INC002"]
        return {
            "year": year,
            "month": month,
            "groups": [
                {
                    "application": "App",
                    "issue": "Test",
                    "incident_numbers": ids,
                    "count": len(ids),
                    "business_impact": "",
                    "recommended_action": "",
                }
            ],
        }

    def test_sets_all_groups(self, tmp_path, monkeypatch):
        monkeypatch.setattr(app, "DATA_DIR", tmp_path)
        load_results_record(self._record())
        assert st.session_state.all_groups is not None
        assert len(st.session_state.all_groups) == 1

    def test_sets_loaded_from_file_true(self, tmp_path, monkeypatch):
        monkeypatch.setattr(app, "DATA_DIR", tmp_path)
        load_results_record(self._record())
        assert st.session_state.loaded_from_file is True

    def test_sets_loaded_year_and_month(self, tmp_path, monkeypatch):
        monkeypatch.setattr(app, "DATA_DIR", tmp_path)
        load_results_record(self._record(year="2024", month="June"))
        assert st.session_state.loaded_year == "2024"
        assert st.session_state.loaded_month == "June"

    def test_clears_workflow_keys(self, tmp_path, monkeypatch):
        monkeypatch.setattr(app, "DATA_DIR", tmp_path)
        st.session_state.batches = [{"some": "batch"}]
        st.session_state.batch_size = 50
        st.session_state.processing_mode = "batch"
        st.session_state.cross_batch_done = True
        load_results_record(self._record())
        assert st.session_state.batches is None
        assert st.session_state.batch_size is None
        assert st.session_state.processing_mode is None
        assert st.session_state.cross_batch_done is None

    def test_loads_companion_csv_when_present(self, tmp_path, monkeypatch):
        monkeypatch.setattr(app, "DATA_DIR", tmp_path)
        csv = tmp_path / "raw_data_2025_January.csv"
        csv.write_text("number,description_raw\nINC001,d1\nINC002,d2\nINC003,d3\n")
        load_results_record(self._record(ids=["INC001", "INC002"]))
        # CSV has 3 rows; df should reflect the CSV, not just the group IDs
        assert len(st.session_state.df) == 3

    def test_synthesises_df_without_csv(self, tmp_path, monkeypatch):
        monkeypatch.setattr(app, "DATA_DIR", tmp_path)
        load_results_record(self._record(ids=["INC001", "INC002"]))
        assert st.session_state.df is not None
        numbers = set(st.session_state.df["number"].astype(str))
        assert numbers == {"INC001", "INC002"}

    def test_synthetic_df_description_raw_placeholder(self, tmp_path, monkeypatch):
        monkeypatch.setattr(app, "DATA_DIR", tmp_path)
        load_results_record(self._record(ids=["INC001"]))
        assert st.session_state.df["description_raw"].iloc[0] == "(not available)"

    def test_computes_missing_ids_from_csv(self, tmp_path, monkeypatch):
        monkeypatch.setattr(app, "DATA_DIR", tmp_path)
        csv = tmp_path / "raw_data_2025_January.csv"
        csv.write_text("number,description_raw\nINC001,d1\nINC002,d2\nINC003,d3\n")
        load_results_record(self._record(ids=["INC001", "INC002"]))
        assert "INC003" in st.session_state.missing_ids

    def test_no_missing_ids_when_all_covered(self, tmp_path, monkeypatch):
        monkeypatch.setattr(app, "DATA_DIR", tmp_path)
        load_results_record(self._record(ids=["INC001", "INC002"]))
        assert st.session_state.missing_ids == []

    def test_assigns_uuids_to_groups_missing_ids(self, tmp_path, monkeypatch):
        monkeypatch.setattr(app, "DATA_DIR", tmp_path)
        record = self._record()
        # Groups in the record have no _id
        load_results_record(record)
        assert "_id" in st.session_state.all_groups[0]

    def test_falls_back_to_synthetic_on_bad_csv_columns(self, tmp_path, monkeypatch):
        monkeypatch.setattr(app, "DATA_DIR", tmp_path)
        csv = tmp_path / "raw_data_2025_January.csv"
        csv.write_text("col_a,col_b\nval1,val2\n")
        load_results_record(self._record(ids=["INC001"]))
        # Should fall back to synthetic df with just the incident IDs
        assert st.session_state.df is not None
        assert len(st.session_state.df) == 1

    def test_handles_empty_groups_list(self, tmp_path, monkeypatch):
        monkeypatch.setattr(app, "DATA_DIR", tmp_path)
        record = {"year": "2025", "month": "January", "groups": []}
        load_results_record(record)
        assert st.session_state.all_groups == []
        assert st.session_state.missing_ids == []

    def test_falls_back_on_csv_read_exception(self, tmp_path, monkeypatch):
        monkeypatch.setattr(app, "DATA_DIR", tmp_path)
        csv = tmp_path / "raw_data_2025_January.csv"
        csv.write_bytes(b"\xff\xfe" + b"garbage binary content")  # unreadable as CSV
        # Should not raise; falls back to synthetic df
        load_results_record(self._record(ids=["INC001"]))
        assert st.session_state.df is not None


# ══════════════════════════════════════════════════════════════════════════════
# update_coverage
# ══════════════════════════════════════════════════════════════════════════════

class TestUpdateCoverage:
    from app import update_coverage

    def test_finds_missing_ids(self):
        from app import update_coverage
        st.session_state.df = pd.DataFrame({
            "number": ["INC001", "INC002", "INC003"],
            "description_raw": ["a", "b", "c"],
        })
        groups = [make_group(gid="g1", ids=["INC001", "INC002"])]
        update_coverage(groups)
        assert st.session_state.missing_ids == ["INC003"]

    def test_no_missing_when_all_covered(self):
        from app import update_coverage
        st.session_state.df = pd.DataFrame({
            "number": ["INC001", "INC002"],
            "description_raw": ["a", "b"],
        })
        groups = [make_group(gid="g1", ids=["INC001", "INC002"])]
        update_coverage(groups)
        assert st.session_state.missing_ids == []

    def test_all_missing_when_no_groups(self):
        from app import update_coverage
        st.session_state.df = pd.DataFrame({
            "number": ["INC001", "INC002"],
            "description_raw": ["a", "b"],
        })
        update_coverage([])
        assert sorted(st.session_state.missing_ids) == ["INC001", "INC002"]

    def test_missing_ids_are_sorted(self):
        from app import update_coverage
        st.session_state.df = pd.DataFrame({
            "number": ["INC003", "INC001", "INC002"],
            "description_raw": ["a", "b", "c"],
        })
        update_coverage([])
        assert st.session_state.missing_ids == sorted(st.session_state.missing_ids)


# ══════════════════════════════════════════════════════════════════════════════
# show_parse_error  (Streamlit calls are mocked — just verify no crash)
# ══════════════════════════════════════════════════════════════════════════════

class TestShowParseError:

    def test_does_not_raise_on_short_text(self):
        from app import show_parse_error
        show_parse_error("bad json")

    def test_does_not_raise_on_long_text(self):
        from app import show_parse_error
        show_parse_error("x" * 1000)

    def test_does_not_raise_on_empty_string(self):
        from app import show_parse_error
        show_parse_error("")


# ══════════════════════════════════════════════════════════════════════════════
# build_excel & _autofit_columns
# ══════════════════════════════════════════════════════════════════════════════

class TestBuildExcel:

    def test_returns_bytes(self):
        from app import build_excel
        result = build_excel([make_group(ids=["INC001"])])
        assert isinstance(result, bytes)
        assert len(result) > 0

    def test_produces_valid_xlsx(self):
        from app import build_excel
        import openpyxl, io
        data = build_excel([make_group(ids=["INC001", "INC002"])])
        wb = openpyxl.load_workbook(io.BytesIO(data))
        assert "Management Summary" in wb.sheetnames
        assert "Incident Detail" in wb.sheetnames

    def test_management_summary_has_data_row(self):
        from app import build_excel
        import openpyxl, io
        groups = [make_group(app_name="Bloomberg", issue="Crash", ids=["INC001"])]
        wb = openpyxl.load_workbook(io.BytesIO(build_excel(groups)))
        ws = wb["Management Summary"]
        # Row 1 = header, row 2 = first data row
        assert ws.cell(row=2, column=1).value == "Bloomberg"

    def test_incident_detail_lists_ids(self):
        from app import build_excel
        import openpyxl, io
        groups = [make_group(ids=["INC001", "INC002"])]
        wb = openpyxl.load_workbook(io.BytesIO(build_excel(groups)))
        ws = wb["Incident Detail"]
        cell_val = ws.cell(row=2, column=4).value
        assert "INC001" in cell_val
        assert "INC002" in cell_val

    def test_creates_unaccounted_sheet_when_df_provided(self):
        from app import build_excel
        import openpyxl, io
        unaccounted = pd.DataFrame({"number": ["INC099"], "description_raw": ["orphan"]})
        wb = openpyxl.load_workbook(io.BytesIO(build_excel([], unaccounted)))
        assert "Unaccounted" in wb.sheetnames

    def test_no_unaccounted_sheet_when_df_is_none(self):
        from app import build_excel
        import openpyxl, io
        wb = openpyxl.load_workbook(io.BytesIO(build_excel([])))
        assert "Unaccounted" not in wb.sheetnames

    def test_no_unaccounted_sheet_when_df_is_empty(self):
        from app import build_excel
        import openpyxl, io
        empty_df = pd.DataFrame({"number": [], "description_raw": []})
        wb = openpyxl.load_workbook(io.BytesIO(build_excel([], empty_df)))
        assert "Unaccounted" not in wb.sheetnames

    def test_groups_sorted_by_count_descending(self):
        from app import build_excel
        import openpyxl, io
        g1 = make_group(app_name="Small", ids=["INC001"])
        g2 = make_group(app_name="Large", ids=["INC002", "INC003", "INC004"])
        wb = openpyxl.load_workbook(io.BytesIO(build_excel([g1, g2])))
        ws = wb["Management Summary"]
        assert ws.cell(row=2, column=1).value == "Large"
