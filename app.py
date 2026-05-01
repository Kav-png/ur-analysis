import re
import json
import io
import uuid
from datetime import datetime
from pathlib import Path

import altair as alt
import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from st_copy import copy_button

st.set_page_config(page_title="Incident Pattern Analyser", layout="wide")

st.title("Incident Pattern Analyser")
st.caption("Surface recurring issues for senior management reporting")

# ── constants ──────────────────────────────────────────────────────────────────
DATA_DIR = Path("data")
now = datetime.now()
months = [
    "January", "February", "March", "April", "May", "June",
    "July", "August", "September", "October", "November", "December"
]
years = [str(y) for y in range(now.year, now.year - 5, -1)]

# ── session state ──────────────────────────────────────────────────────────────
for key in ("df", "batches", "batch_size", "all_groups", "missing_ids",
            "processing_mode", "cross_batch_done",
            "loaded_from_file", "loaded_year", "loaded_month"):
    if key not in st.session_state:
        st.session_state[key] = None
if "merge_selected" not in st.session_state:
    st.session_state.merge_selected = set()


# ── helpers ────────────────────────────────────────────────────────────────────
def preprocess(text: str, maxlen: int = 500) -> str:
    text = str(text)
    # remove ticket/reference IDs (e.g. INC12345, CHG99999)
    text = re.sub(r'\b[A-Z]{2,4}\d{5,}\b', '', text)
    # remove email addresses
    text = re.sub(r'\S+@\S+', '', text)
    # remove phone numbers (must look like a phone: starts with + or has spaces/dashes between digits)
    text = re.sub(r'Tel\.?[\s:\-]*\d+', '', text, flags=re.IGNORECASE)
    text = re.sub(r'\+\d[\d\s\-]{7,}', '', text)
    # remove boilerplate / disclaimer lines
    text = re.sub(r'Client Identifying Data.*?not allowed.*?\.?', '', text, flags=re.IGNORECASE)
    text = re.sub(r'No CID Disclaimer accepted: TRUE', '', text, flags=re.IGNORECASE)
    text = re.sub(r'I have read and understood the disclaimer.*', '', text, flags=re.IGNORECASE)
    # remove generic salutations and sign-offs
    text = re.sub(r'^(hi|hello|dear|greetings)[\s,]+', '', text, flags=re.IGNORECASE)
    text = re.sub(r'(thanks|thank you|regards|best regards|cheers)[\s,]*$', '', text, flags=re.IGNORECASE)
    # remove user detail blocks
    text = re.sub(r'User details:.*?(?=Topic:|$)', '', text, flags=re.IGNORECASE)
    # remove internal tracker tokens like (T12345)
    text = re.sub(r'\(T\d+\)', '', text)
    # collapse separators and whitespace
    text = re.sub(r'[-_]{2,}', ' ', text)
    text = re.sub(r'\s+', ' ', text).strip()
    return text[:maxlen]


def build_prompt(rows: list[dict]) -> str:
    incidents = "\n".join(
        f"{i+1}. [{r['number']}] {r['description']}" for i, r in enumerate(rows)
    )
    return f"""You are analyzing {len(rows)} IT incidents from an Investment Bank.
Your output will be used by senior management to reduce recurring incidents.

## Step 1 — Scratchpad (plain text, not JSON)
Before producing any JSON, reason through the incidents:
- List the distinct root-cause themes or recurring request types you can see.
- Note any incidents that are ambiguous or hard to classify.
- Decide which incidents belong together and why.
This thinking is for your benefit only; it will be discarded.

## Step 2 — Verify coverage
Count the incidents you intend to include in your JSON. It must equal exactly {len(rows)}.
If any are missing, add them before continuing.

## Step 3 — Output JSON
Rules:
- Extract the application or system name from the description text.
- Group incidents together if they share the same root cause OR if they represent repeated requests for the same operational process or workflow (e.g., repeated requests to create Storm IDs, repeated manual report retriggers, etc.).
- Be specific: write "Users locked out after AD password reset" not "access issue".
- If only one incident exists for a particular issue, include it as its own group (do not omit unique incidents).
- Every incident must appear in the output, either as part of a group or as its own group. Do not omit any incident for any reason.
- If the application cannot be identified from the text, use "Unknown System".
- After the scratchpad, return the JSON block below and nothing else.

Output format:
{{
  "groups": [
    {{
      "application": "name of the system or application",
      "issue": "precise description of the problem or request",
      "count": <number>,
      "incident_numbers": ["INC001", "INC002"],
      "business_impact": "one sentence describing the business effect",
      "recommended_action": "one concrete action to prevent recurrence or address the issue"
    }}
  ]
}}

Incidents:
{incidents}
"""


def ensure_group_ids(groups: list[dict]) -> list[dict]:
    for g in groups:
        if "_id" not in g:
            g["_id"] = str(uuid.uuid4())
    return groups


def normalise_quotes(text: str) -> str:
    return (text
        .replace('“', '"').replace('”', '"')
        .replace('‘', "'").replace('’', "'")
        .replace('′', "'").replace('″', '"')
        .replace('﻿', '')
    )


def parse_response(raw: str) -> list[dict] | None:
    cleaned = re.sub(r"```json|```", "", raw).strip()
    cleaned = normalise_quotes(cleaned)
    json_match = re.search(r'\{.*\}', cleaned, re.DOTALL)
    if json_match:
        cleaned = json_match.group(0)
    try:
        data = json.loads(cleaned)
        raw_groups = data.get("groups") or data.get("g", [])
    except json.JSONDecodeError:
        return None

    groups = []
    seen: set[str] = set()
    for g in raw_groups:
        ids_raw = g.get("incident_numbers") or g.get("ids", [])
        unique_ids = []
        for n in ids_raw:
            key = str(n).strip()
            if key and key not in seen:
                seen.add(key)
                unique_ids.append(key)
        if not unique_ids:
            continue
        groups.append({
            "_id":                str(uuid.uuid4()),
            "application":        g.get("application") or g.get("app", "Unknown System"),
            "issue":              g.get("issue") or g.get("iss", ""),
            "incident_numbers":   unique_ids,
            "count":              len(unique_ids),
            "business_impact":    g.get("business_impact") or g.get("imp", ""),
            "recommended_action": g.get("recommended_action") or g.get("act", ""),
        })
    return groups


def build_cross_batch_prompt(groups: list[dict]) -> str:
    numbered = "\n".join(
        f"{i+1}. [{g.get('application','?')}] {g.get('issue','')}  (count: {g.get('count',0)})"
        for i, g in enumerate(groups)
    )
    return f"""You are a senior IT incident analyst reviewing grouped incident data from an Investment Bank.
The groups below were produced by analysing batches of incidents separately. Similar issues may have
been described slightly differently and placed into distinct groups when they should be one.

## Step 1 — Scratchpad (plain text, not JSON)
Before producing any JSON, reason through the groups:
- Identify clusters of groups that likely share the same root cause, even if worded differently.
- Note any borderline cases and decide — when in doubt, prefer merging over splitting.
- Confirm every group number from 1 to {len(groups)} is accounted for.

## Step 2 — Output JSON
Rules:
- For each set of groups to merge, produce one entry with source_indices listing their 1-based numbers.
- Groups that are genuinely distinct appear on their own with a single-element source_indices.
- Every group number from 1 to {len(groups)} must appear in exactly one source_indices list.
- Do NOT reproduce incident numbers — Python will handle the merge.
- After the scratchpad, return the JSON block below and nothing else.

Output format:
{{
  "groups": [
    {{
      "application": "name of the system or application",
      "issue": "precise description of the problem or request",
      "business_impact": "one sentence describing the business effect",
      "recommended_action": "one concrete action to prevent recurrence or address the issue",
      "source_indices": [1, 4, 7]
    }}
  ]
}}

Groups to consolidate:
{numbered}
"""


def apply_cross_batch_merge(original_groups: list[dict], merge_spec: list[dict]) -> list[dict] | None:
    n = len(original_groups)
    assigned = [False] * n
    result = []
    for spec in merge_spec:
        indices = spec.get("source_indices") or []
        zero_based = [i - 1 for i in indices if isinstance(i, int) and 1 <= i <= n]
        if not zero_based:
            return None
        incident_numbers: list[str] = []
        for idx in zero_based:
            incident_numbers.extend(original_groups[idx].get("incident_numbers", []))
            assigned[idx] = True
        result.append({
            "_id":                str(uuid.uuid4()),
            "application":        spec.get("application", "Unknown System"),
            "issue":              spec.get("issue", ""),
            "incident_numbers":   incident_numbers,
            "count":              len(incident_numbers),
            "business_impact":    spec.get("business_impact", ""),
            "recommended_action": spec.get("recommended_action", ""),
        })
    if not all(assigned):
        return None
    return result



def merge_all_batches(batches: list[dict]) -> list[dict]:
    seen: set[str] = set()
    merged: list[dict] = []
    for batch in batches:
        for g in batch.get("groups", []):
            unique_ids = [n for n in g["incident_numbers"] if n not in seen]
            seen.update(unique_ids)
            if unique_ids:
                existing = next(
                    (m for m in merged
                     if m["application"].lower() == g["application"].lower()
                     and m["issue"].lower() == g["issue"].lower()),
                    None
                )
                if existing:
                    existing["incident_numbers"].extend(unique_ids)
                    existing["count"] = len(existing["incident_numbers"])
                else:
                    merged.append({**g, "_id": str(uuid.uuid4()),
                                   "incident_numbers": unique_ids, "count": len(unique_ids)})
    return merged


def _autofit_columns(ws, min_width: int = 10, max_width: int = 60) -> None:
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        best = min_width
        for cell in col_cells:
            if cell.value is not None:
                longest_line = max(len(str(line)) for line in str(cell.value).splitlines() or [""])
                best = max(best, longest_line)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.column_dimensions[col_letter].width = min(best + 2, max_width)


def build_excel(groups: list[dict], unaccounted_df: pd.DataFrame | None = None) -> bytes:
    wb = Workbook()
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F3864")
    alt_fill    = PatternFill("solid", fgColor="DCE6F1")

    ws1 = wb.active
    ws1.title = "Management Summary"
    headers = ["Application", "Issue", "Count", "Business Impact", "Recommended Action"]
    ws1.append(headers)
    for col_idx, _ in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    ws1.freeze_panes = "A2"
    sorted_groups = sorted(groups, key=lambda g: g.get("count", 0), reverse=True)
    for row_idx, g in enumerate(sorted_groups, 2):
        ws1.append([g.get("application",""), g.get("issue",""), g.get("count",0),
                    g.get("business_impact",""), g.get("recommended_action","")])
        if row_idx % 2 == 0:
            for col_idx in range(1, len(headers)+1):
                ws1.cell(row=row_idx, column=col_idx).fill = alt_fill
    _autofit_columns(ws1)

    ws2 = wb.create_sheet("Incident Detail")
    detail_headers = ["Group", "Application", "Issue", "Incident Numbers"]
    ws2.append(detail_headers)
    for col_idx, _ in enumerate(detail_headers, 1):
        cell = ws2.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    ws2.freeze_panes = "A2"
    for group_idx, g in enumerate(sorted_groups, 1):
        ws2.append([group_idx, g.get("application",""), g.get("issue",""),
                    ", ".join(g.get("incident_numbers",[]))])
    _autofit_columns(ws2)

    if unaccounted_df is not None and not unaccounted_df.empty:
        ws3 = wb.create_sheet("Unaccounted")
        ws3.append(["Incident Number", "Description"])
        for col_idx in range(1, 3):
            cell = ws3.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = PatternFill("solid", fgColor="C00000")
            cell.alignment = Alignment(horizontal="center")
        ws3.freeze_panes = "A2"
        for _, row in unaccounted_df.iterrows():
            ws3.append([str(row["number"]), str(row["description_raw"])])
        _autofit_columns(ws3)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def save_monthly_data(df: pd.DataFrame, groups: list[dict], year: str, month: str) -> tuple[str, str]:
    DATA_DIR.mkdir(exist_ok=True)
    raw_path  = DATA_DIR / f"raw_data_{year}_{month}.csv"
    proc_path = DATA_DIR / f"processed_data_{year}_{month}.json"
    df[["number", "description_raw"]].to_csv(raw_path, index=False)
    proc_path.write_text(json.dumps({"year": year, "month": month, "groups": groups}, indent=2))
    return str(raw_path), str(proc_path)


def load_history() -> list[dict]:
    if not DATA_DIR.exists():
        return []
    records = []
    for f in sorted(DATA_DIR.glob("processed_data_*.json")):
        try:
            records.append(json.loads(f.read_text()))
        except (json.JSONDecodeError, OSError):
            pass
    return records


def load_results_record(record: dict) -> None:
    year   = record.get("year", "")
    month  = record.get("month", "")
    groups = record.get("groups", [])

    csv_path = DATA_DIR / f"raw_data_{year}_{month}.csv"
    df = None
    if csv_path.exists():
        try:
            df = pd.read_csv(csv_path)
            df.columns = [c.lower() for c in df.columns]
            if "number" not in df.columns or "description_raw" not in df.columns:
                df = None
        except Exception:
            df = None

    if df is None:
        all_ids = [n for g in groups for n in g.get("incident_numbers", [])]
        df = pd.DataFrame({"number": all_ids, "description_raw": ["(not available)"] * len(all_ids)})

    all_ids_set = set(df["number"].astype(str))
    covered     = {n for g in groups for n in g.get("incident_numbers", [])}
    ensure_group_ids(groups)

    st.session_state.df               = df
    st.session_state.all_groups       = groups
    st.session_state.missing_ids      = sorted(all_ids_set - covered)
    st.session_state.loaded_from_file = True
    st.session_state.loaded_year      = year
    st.session_state.loaded_month     = month
    st.session_state.batches          = None
    st.session_state.batch_size       = None
    st.session_state.processing_mode  = None
    st.session_state.cross_batch_done = None


def red_portal_button() -> None:
    st.markdown(
        """<a href="https://goto/red" target="_blank">
            <button style="background-color:red;color:white;padding:0.5em 1.5em;
                border:none;border-radius:4px;font-size:1em;cursor:pointer;">
                Go to Red Portal
            </button>
        </a>""",
        unsafe_allow_html=True,
    )


def prompt_panel(prompt: str, copy_label: str, key: str, height: int = 200,
                 placeholder: str = '{"groups": [...]}') -> str:
    st.code("\n".join(prompt.splitlines()[:10]) + "\n...", language=None)
    with st.expander("Show full prompt"):
        st.code(prompt, language=None)
    copy_button(prompt, tooltip=copy_label, copied_label="Copied!", icon="st")
    red_portal_button()
    return st.text_area("Paste AI response here", height=height, key=key, placeholder=placeholder)


def show_parse_error(raw: str) -> None:
    st.error("Could not parse as JSON.")
    with st.expander("Debug — show received text"):
        st.text(f"First 300 chars:\n{raw[:300]}")
        st.text(f"Last 100 chars:\n{raw[-100:]}")


def update_coverage(groups: list[dict]) -> None:
    all_ids = set(st.session_state.df["number"].astype(str))
    covered = {n for g in groups for n in g["incident_numbers"]}
    st.session_state.missing_ids = sorted(all_ids - covered)


def _find_group(gid: str) -> dict | None:
    return next((g for g in (st.session_state.all_groups or []) if g["_id"] == gid), None)


def _sync_field(gid: str, field: str, widget_key: str) -> None:
    g = _find_group(gid)
    if g is not None:
        g[field] = st.session_state[widget_key]


def _delete_group(gid: str) -> None:
    g = _find_group(gid)
    if g is None:
        return
    freed = g.get("incident_numbers", [])
    st.session_state.all_groups = [x for x in st.session_state.all_groups if x["_id"] != gid]
    st.session_state.missing_ids = sorted(
        set(st.session_state.missing_ids or []) | set(freed)
    )
    st.session_state.merge_selected.discard(gid)


def _add_to_group(gid: str, ids_to_add: list[str]) -> None:
    g = _find_group(gid)
    if g is None or not ids_to_add:
        return
    g["incident_numbers"] = list(dict.fromkeys(g["incident_numbers"] + ids_to_add))
    g["count"] = len(g["incident_numbers"])
    st.session_state.missing_ids = sorted(
        set(st.session_state.missing_ids or []) - set(ids_to_add)
    )


def _merge_groups(gids: list[str], new_issue: str) -> None:
    sources = [g for g in st.session_state.all_groups if g["_id"] in gids]
    if len(sources) < 2:
        return
    merged_ids: list[str] = []
    seen: set[str] = set()
    for g in sources:
        for n in g["incident_numbers"]:
            if n not in seen:
                seen.add(n)
                merged_ids.append(n)
    new_group = {
        "_id":                str(uuid.uuid4()),
        "application":        sources[0]["application"],
        "issue":              new_issue or sources[0]["issue"],
        "incident_numbers":   merged_ids,
        "count":              len(merged_ids),
        "business_impact":    sources[0]["business_impact"],
        "recommended_action": sources[0]["recommended_action"],
    }
    keep = [g for g in st.session_state.all_groups if g["_id"] not in gids]
    st.session_state.all_groups = keep + [new_group]
    st.session_state.merge_selected = set()



# ══════════════════════════════════════════════════════════════════════════════
# SIDEBAR — Browse & Load Saved Reports
# ══════════════════════════════════════════════════════════════════════════════
with st.sidebar:
    st.header("Saved Reports")

    _history = load_history()
    if _history:
        _labels = [f"{r['month']} {r['year']}" for r in _history]
        _choice = st.selectbox("Select a saved report", ["— choose —"] + _labels,
                               key="sidebar_select")
        if st.button("Load Selected", key="sidebar_load_btn",
                     disabled=(_choice == "— choose —")):
            load_results_record(_history[_labels.index(_choice)])
            st.rerun()
    else:
        st.info("No saved reports found in `data/`.")

    st.divider()
    st.subheader("Upload a JSON Report")
    _uploaded_json = st.file_uploader("Upload processed_data JSON", type="json",
                                      key="sidebar_json_upload")
    if _uploaded_json is not None:
        try:
            _record = json.loads(_uploaded_json.read())
            if st.button("Load Uploaded JSON", key="sidebar_json_load_btn"):
                load_results_record(_record)
                st.rerun()
        except (json.JSONDecodeError, OSError):
            st.error("Invalid JSON file — could not parse.")

    if st.session_state.loaded_from_file:
        st.divider()
        if st.button("Clear loaded report", key="sidebar_clear_btn", type="secondary"):
            for _k in ("df", "batches", "batch_size", "all_groups", "missing_ids",
                       "processing_mode", "cross_batch_done",
                       "loaded_from_file", "loaded_year", "loaded_month"):
                st.session_state[_k] = None
            st.rerun()


# ══════════════════════════════════════════════════════════════════════════════
# STEP 1 — Upload Excel
# ══════════════════════════════════════════════════════════════════════════════
if not st.session_state.loaded_from_file:
    st.header("Step 1 — Upload Incident Excel")

    uploaded = st.file_uploader("Upload your monthly incidents Excel file", type="xlsx")
    if uploaded:
        df_raw = pd.read_excel(uploaded)
        cols = df_raw.columns.tolist()

        col1, col2 = st.columns(2)
        num_col = col1.selectbox(
            "Incident number column", cols,
            index=next((i for i, c in enumerate(cols)
                        if "number" in c.lower() or "id" in c.lower() or "inc" in c.lower()), 0)
        )
        desc_col = col2.selectbox(
            "Description column", cols,
            index=(next((i for i, c in enumerate(cols) if c.lower() == "description"),
                   next((i for i, c in enumerate(cols)
                         if "description" in c.lower() or "summary" in c.lower()), min(1, len(cols)-1))))
        )

        maxlen = st.slider(
            "Max description length after preprocessing (characters)",
            min_value=100, max_value=2000, value=500, step=50,
            help="Increase if important context is being cut off. The preview below updates live.",
        )

        df = df_raw[[num_col, desc_col]].dropna(subset=[desc_col]).copy()
        df.columns = ["number", "description_raw"]
        df["description"] = df["description_raw"].apply(lambda x: preprocess(x, maxlen=maxlen))
        st.session_state.df = df

        st.success(f"{len(df)} incidents loaded")

        with st.expander("Preview — Before & After Preprocessing (first 5 rows)", expanded=True):
            preview = df[["number", "description_raw", "description"]].head(5).copy()
            preview.columns = ["Incident Number", "Description (Before)", "Description (After)"]
            st.dataframe(preview, use_container_width=True, hide_index=True)


# ══════════════════════════════════════════════════════════════════════════════
# STEP 2 — Choose Processing Mode & Generate Prompts
# ══════════════════════════════════════════════════════════════════════════════
if st.session_state.df is not None and not st.session_state.loaded_from_file:
    st.divider()
    st.header("Step 2 — Choose Processing Mode")

    mode = st.radio(
        "How do you want to analyse the incidents?",
        options=["batch", "all"],
        format_func=lambda x: (
            "Batch processing — split into smaller chunks (recommended for large datasets)"
            if x == "batch"
            else "Process all incidents together — single prompt (best for smaller datasets)"
        ),
        index=0 if st.session_state.processing_mode != "all" else 1,
        horizontal=True,
        key="mode_radio",
    )
    st.session_state.processing_mode = mode

    use_preprocess = st.toggle(
        "Apply preprocessing to descriptions before generating prompts",
        value=True,
        help="When on, ticket IDs, emails, phone numbers, and boilerplate are stripped before sending to the AI. Turn off to send raw descriptions.",
    )
    desc_col = "description" if use_preprocess else "description_raw"

    total = len(st.session_state.df)

    # ── Option A: All together ─────────────────────────────────────────────────
    if mode == "all":
        st.subheader("Step 2A — Process All Incidents Together")
        st.info(f"This generates a single prompt for all {total} incidents. Best for datasets under ~100 incidents.")

        rows = st.session_state.df[["number", desc_col]].rename(columns={desc_col: "description"}).to_dict("records")
        full_prompt = build_prompt(rows)

        raw_all = prompt_panel(full_prompt, "Copy full prompt", "response_all")

        if st.button("Process All Incidents", type="primary", key="process_all"):
            if not raw_all.strip():
                st.warning("Paste the AI response before processing.")
            else:
                groups = parse_response(raw_all)
                if groups is None:
                    show_parse_error(raw_all)
                elif len(groups) == 0:
                    st.warning("No groups found — try re-running the prompt.")
                else:
                    st.session_state.batches = [{
                        "index": 0, "rows": rows, "prompt": full_prompt,
                        "groups": groups, "complete": True,
                    }]
                    st.session_state.batch_size = len(rows)
                    st.session_state.all_groups = groups
                    update_coverage(groups)
                    st.session_state.cross_batch_done = True
                    st.rerun()

        if st.session_state.all_groups and st.session_state.processing_mode == "all":
            st.success(f"Processed — {len(st.session_state.all_groups)} groups found. Continue to Step 3 below.")

    # ── Option B: Batch processing ─────────────────────────────────────────────
    else:
        st.subheader("Step 2B — Configure & Analyse Batches")

        batch_size = st.slider(
            "Incidents per batch", min_value=25, max_value=150, value=75, step=25,
            help="Smaller batches = more paste-backs but better AI coverage per batch. 50–75 is the sweet spot."
        )

        n_batches = -(-total // batch_size)
        st.caption(f"{total} incidents → **{n_batches} batch{'es' if n_batches != 1 else ''}** of up to {batch_size}")

        if st.button("Generate Batch Prompts", type="primary"):
            rows = st.session_state.df[["number", desc_col]].rename(columns={desc_col: "description"}).to_dict("records")
            st.session_state.batches = [
                {"index": i, "rows": chunk, "prompt": build_prompt(chunk), "groups": None, "complete": False}
                for i, chunk in (
                    (i, rows[i * batch_size: (i + 1) * batch_size]) for i in range(n_batches)
                )
            ]
            st.session_state.batch_size    = batch_size
            st.session_state.all_groups    = None
            st.session_state.missing_ids   = None
            st.session_state.cross_batch_done = None

        if st.session_state.batches:
            batches  = st.session_state.batches
            completed = sum(1 for b in batches if b["complete"])
            all_done  = completed == len(batches)

            st.progress(completed / len(batches), text=f"{completed}/{len(batches)} batches complete")

            # ── Completed summaries (collapsed) ───────────────────────────────
            done_batches = [b for b in batches if b["complete"]]
            if done_batches:
                with st.expander(f"Completed batches ({len(done_batches)})", expanded=False):
                    for b in done_batches:
                        start = b['index'] * st.session_state.batch_size + 1
                        end   = start + len(b['rows']) - 1
                        n_groups = len(b['groups'])
                        n_inc    = sum(g['count'] for g in b['groups'])
                        st.success(
                            f"Batch {b['index']+1} (incidents {start}–{end}) — "
                            f"{n_inc} incidents → {n_groups} groups",
                            icon="✅",
                        )
                        if st.button(f"Re-do batch {b['index']+1}", key=f"redo_{b['index']}",
                                     type="secondary"):
                            st.session_state.batches[b['index']]["complete"] = False
                            st.session_state.batches[b['index']]["groups"]   = None
                            st.session_state.cross_batch_done = None
                            st.rerun()

            # ── Active batch: first incomplete one ────────────────────────────
            active = next((b for b in batches if not b["complete"]), None)
            if active is not None:
                start = active['index'] * st.session_state.batch_size + 1
                end   = start + len(active['rows']) - 1
                st.subheader(f"Batch {active['index']+1} of {len(batches)}  —  incidents {start}–{end}")
                st.info("Copy the prompt, run it in your AI tool, then paste the response below.")

                raw_response = prompt_panel(
                    active["prompt"],
                    f"Copy Batch {active['index']+1} prompt",
                    f"response_batch_{active['index']}",
                )

                if st.button(f"Process Batch {active['index']+1}", type="primary",
                             key=f"process_{active['index']}"):
                    if not raw_response.strip():
                        st.warning("Paste the AI response before processing.")
                    else:
                        groups = parse_response(raw_response)
                        if groups is None:
                            show_parse_error(raw_response)
                        elif len(groups) == 0:
                            st.warning("No groups found — try re-running the prompt.")
                        else:
                            st.session_state.batches[active["index"]]["groups"]   = groups
                            st.session_state.batches[active["index"]]["complete"] = True
                            st.session_state.all_groups = merge_all_batches(st.session_state.batches)
                            update_coverage(st.session_state.all_groups)
                            st.session_state.cross_batch_done = False
                            st.rerun()

            # ── Cross-batch consolidation (inline, only when all batches done) ─
            elif all_done and len(batches) > 1:
                st.divider()
                if st.session_state.cross_batch_done:
                    st.success(f"Cross-batch consolidation complete — {len(st.session_state.all_groups)} final groups.")
                    if st.button("Re-run consolidation", type="secondary", key="rerun_crossbatch"):
                        st.session_state.cross_batch_done = False
                        st.rerun()
                else:
                    st.subheader("Final step — Consolidate across batches")
                    st.info(
                        f"All {len(batches)} batches done ({len(st.session_state.all_groups)} groups so far). "
                        "Copy the prompt below so the AI can merge any duplicates across batches."
                    )
                    raw_cross = prompt_panel(
                        build_cross_batch_prompt(st.session_state.all_groups),
                        "Copy consolidation prompt",
                        "response_cross_batch",
                        height=250,
                    )
                    if st.button("Process Consolidation Response", type="primary", key="process_cross_batch"):
                        if not raw_cross.strip():
                            st.warning("Paste the AI response before processing.")
                        else:
                            cleaned = re.sub(r"```json|```", "", raw_cross).strip()
                            cleaned = normalise_quotes(cleaned)
                            match = re.search(r'\{.*\}', cleaned, re.DOTALL)
                            merge_spec = None
                            if match:
                                try:
                                    merge_spec = json.loads(match.group(0)).get("groups", [])
                                except json.JSONDecodeError:
                                    pass
                            if merge_spec is None:
                                show_parse_error(raw_cross)
                            else:
                                refined_groups = apply_cross_batch_merge(
                                    st.session_state.all_groups, merge_spec
                                )
                                if refined_groups is None:
                                    st.error(
                                        "Consolidation failed — the AI's source_indices didn't cover "
                                        "all groups or referenced invalid numbers. Try re-running the prompt."
                                    )
                                    with st.expander("Debug — show received text"):
                                        st.text(raw_cross[:500])
                                elif len(refined_groups) == 0:
                                    st.warning("No groups returned — try re-running the prompt.")
                                else:
                                    st.session_state.all_groups = refined_groups
                                    update_coverage(refined_groups)
                                    st.session_state.cross_batch_done = True
                                    st.rerun()


# ══════════════════════════════════════════════════════════════════════════════
# STEP 4 / 3 — Results & Export
# ══════════════════════════════════════════════════════════════════════════════
if st.session_state.all_groups:
    st.divider()
    st.header("Step 3 — Results")

    groups       = st.session_state.all_groups
    missing_ids  = st.session_state.missing_ids or []
    batches_done = sum(1 for b in (st.session_state.batches or []) if b["complete"])
    batches_total= len(st.session_state.batches or [])

    col_month, col_year = st.columns([2, 1])
    _mi = months.index(st.session_state.loaded_month) if st.session_state.loaded_month in months else now.month - 1
    _yi = years.index(st.session_state.loaded_year)   if st.session_state.loaded_year  in years  else 0
    selected_month = col_month.selectbox("Report Month", months, index=_mi)
    selected_year  = col_year.selectbox("Report Year",   years,  index=_yi)

    total_covered = sum(g.get("count", 0) for g in groups)
    applications  = sorted({g.get("application", "Unknown") for g in groups})

    m0, m1, m2, m3, m4 = st.columns(5)
    m0.metric("Total Incidents",   len(st.session_state.df))
    m1.metric("Batches Complete",
              "Loaded from file" if st.session_state.loaded_from_file
              else f"{batches_done}/{batches_total}")
    m2.metric("Incidents Covered", total_covered)
    m3.metric("Groups Found",      len(groups))
    m4.metric("Applications",      len(applications))

    if missing_ids:
        st.warning(f"⚠ {len(missing_ids)} incidents not yet covered — complete remaining batches or check AI responses.")
        with st.expander(f"Show {len(missing_ids)} uncovered incidents"):
            missing_df = st.session_state.df[st.session_state.df["number"].astype(str).isin(missing_ids)]
            st.dataframe(missing_df[["number", "description_raw"]], use_container_width=True, hide_index=True)
    else:
        st.success(
            f"All {len(st.session_state.df)} incidents covered."
            if st.session_state.loaded_from_file
            else f"All {len(st.session_state.df)} incidents covered across {batches_done} batches."
        )

    # ── Donut chart — incidents by application ────────────────────────────────
    app_totals = {}
    for g in groups:
        app = g.get("application", "Unknown")
        app_totals[app] = app_totals.get(app, 0) + g.get("count", 0)
    donut_df = pd.DataFrame(
        [{"application": a, "count": c} for a, c in sorted(app_totals.items(), key=lambda x: -x[1])]
    )
    donut = (
        alt.Chart(donut_df)
        .mark_arc(innerRadius=80, outerRadius=160)
        .encode(
            theta=alt.Theta("count:Q"),
            color=alt.Color(
                "application:N",
                legend=alt.Legend(title="Application", orient="right"),
            ),
            tooltip=[
                alt.Tooltip("application:N", title="Application"),
                alt.Tooltip("count:Q",       title="Incidents"),
            ],
        )
        .properties(
            title=f"{selected_month} {selected_year} — Incidents by Application",
            height=380,
        )
    )
    st.altair_chart(donut, use_container_width=True)

    st.subheader("Groups by Application")

    ensure_group_ids(groups)
    sorted_groups = sorted(groups, key=lambda g: g.get("count", 0), reverse=True)

    # ── Merge bar (appears when 2+ groups are checked) ───────────────────────
    sel = st.session_state.merge_selected
    if len(sel) >= 2:
        with st.container(border=True):
            mc1, mc2 = st.columns([3, 1])
            mc1.markdown(f"**{len(sel)} groups selected for merge**")
            merge_issue = mc1.text_input(
                "Merged issue description",
                placeholder="Write a description for the merged group…",
                key="merge_issue_input",
                label_visibility="collapsed",
            )
            if mc2.button("Merge", type="primary", key="do_merge_btn"):
                _merge_groups(list(sel), merge_issue)
                groups = st.session_state.all_groups
                update_coverage(groups)
                st.rerun()
            if mc2.button("Clear selection", key="clear_merge_btn"):
                st.session_state.merge_selected = set()
                st.rerun()

    # ── Group expanders ───────────────────────────────────────────────────────
    for app in sorted({g.get("application", "Unknown") for g in sorted_groups},
                      key=lambda a: sum(g["count"] for g in sorted_groups if g.get("application") == a),
                      reverse=True):
        app_groups = [g for g in sorted_groups if g.get("application") == app]
        app_total  = sum(g.get("count", 0) for g in app_groups)
        with st.expander(f"{app}  —  {app_total} incident{'s' if app_total != 1 else ''}"):
            for g in app_groups:
                gid = g["_id"]

                # Merge checkbox
                checked = st.checkbox(
                    "Select for merge", value=(gid in sel),
                    key=f"chk_{gid}",
                )
                if checked:
                    st.session_state.merge_selected.add(gid)
                else:
                    st.session_state.merge_selected.discard(gid)

                # Editable issue
                issue_key = f"issue_{gid}"
                if issue_key not in st.session_state:
                    st.session_state[issue_key] = g.get("issue", "")
                st.text_area(
                    "Issue",
                    key=issue_key,
                    on_change=_sync_field,
                    args=(gid, "issue", issue_key),
                    height=80,
                )

                col_count, col_imp, col_act = st.columns([1, 3, 3])
                col_count.metric("Count", g.get("count", 0))

                # Editable business impact
                imp_key = f"imp_{gid}"
                if imp_key not in st.session_state:
                    st.session_state[imp_key] = g.get("business_impact", "")
                col_imp.text_area(
                    "Business Impact",
                    key=imp_key,
                    on_change=_sync_field,
                    args=(gid, "business_impact", imp_key),
                    height=100,
                )

                # Editable recommended action
                act_key = f"act_{gid}"
                if act_key not in st.session_state:
                    st.session_state[act_key] = g.get("recommended_action", "")
                col_act.text_area(
                    "Recommended Action",
                    key=act_key,
                    on_change=_sync_field,
                    args=(gid, "recommended_action", act_key),
                    height=100,
                )

                # Incident numbers
                inc_nums = g.get("incident_numbers", [])
                if inc_nums:
                    st.caption("Incidents: " + " · ".join(inc_nums))

                # Add unaccounted incidents
                current_missing = st.session_state.missing_ids or []
                if current_missing:
                    add_sel = st.multiselect(
                        "Add unaccounted incidents to this group",
                        options=current_missing,
                        key=f"add_{gid}",
                        placeholder="Select incidents to add…",
                    )
                    if add_sel and st.button("Add to group", key=f"addbtn_{gid}"):
                        _add_to_group(gid, add_sel)
                        st.session_state.pop(f"add_{gid}", None)
                        update_coverage(st.session_state.all_groups)
                        st.rerun()

                # Delete group
                if st.button("Delete group", key=f"del_{gid}", type="secondary"):
                    _delete_group(gid)
                    update_coverage(st.session_state.all_groups)
                    st.rerun()

                st.divider()

    # ── Save ──
    if not st.session_state.loaded_from_file:
        st.subheader("Save Monthly Data")
        if st.button(f"Save {selected_month} {selected_year} Data", type="secondary"):
            raw_path, proc_path = save_monthly_data(st.session_state.df, groups, selected_year, selected_month)
            st.success(f"Saved:\n- `{raw_path}`\n- `{proc_path}`")

    # ── Excel Export ──
    st.subheader("Export")
    unaccounted_df = (
        st.session_state.df[st.session_state.df["number"].astype(str).isin(missing_ids)]
        if missing_ids else None
    )
    excel_bytes    = build_excel(groups, unaccounted_df)
    excel_filename = f"universal_request_incident_patterns_{selected_year}-{selected_month}.xlsx"
    st.download_button(
        label="Download Excel Report",
        data=excel_bytes,
        file_name=excel_filename,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary",
    )

    # ── History & Trends ──
    st.divider()
    st.subheader("Monthly Trends")
    history = load_history()

    if len(history) < 2:
        st.info("Save data for 2 or more months to see trend analysis.")
    else:
        month_idx = {m: i for i, m in enumerate(months)}
        sorted_history = sorted(history, key=lambda r: (int(r["year"]), month_idx.get(r["month"], 0)))

        month_order = [f"{r['month'][:3]} {r['year']}" for r in sorted_history]

        hist_df = pd.DataFrame([
            {"month": f"{r['month'][:3]} {r['year']}",
             "application": g.get("application", "Unknown"),
             "count": g.get("count", 0)}
            for r in sorted_history for g in r.get("groups", [])
        ])

        df_vol = pd.DataFrame([
            {"month": f"{r['month'][:3]} {r['year']}",
             "total": sum(g.get("count", 0) for g in r.get("groups", []))}
            for r in sorted_history
        ])
        df_vol["moving_avg"] = df_vol["total"].rolling(window=3, min_periods=1).mean().round(1)
        baseline             = df_vol["total"].iloc[0]

        target_pct = st.slider(
            "Target reduction per month (%)", min_value=1, max_value=30, value=10, step=1
        )
        df_vol["target"] = [round(baseline * ((1 - target_pct / 100) ** i), 1) for i in range(len(df_vol))]

        bars = alt.Chart(hist_df).mark_bar(opacity=0.9).encode(
            x=alt.X("month:N", sort=month_order, title="Month", axis=alt.Axis(labelAngle=-30)),
            y=alt.Y("sum(count):Q", title="Total Incidents"),
            color=alt.Color("application:N", legend=alt.Legend(title="Application")),
            tooltip=[
                alt.Tooltip("month:N",          title="Month"),
                alt.Tooltip("application:N",     title="Application"),
                alt.Tooltip("count:Q",           title="Incidents"),
            ],
        )
        base_vol = alt.Chart(df_vol).encode(
            x=alt.X("month:N", sort=month_order)
        )
        ma_line     = base_vol.mark_line(color="#F4A900", strokeWidth=2.5, point=True).encode(y="moving_avg:Q")
        target_line = base_vol.mark_line(color="#C00000", strokeWidth=2, strokeDash=[6, 3]).encode(y="target:Q")

        st.altair_chart(
            (bars + ma_line + target_line).properties(height=380).configure_axis(labelFontSize=12, titleFontSize=13),
            use_container_width=True
        )
        st.caption("Orange line = 3-month moving average  |  Red dashed = target reduction trajectory")


